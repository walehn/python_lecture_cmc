import pandas as pd
import glob, os
import tkinter as tk
import tkinter.ttk
from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import datetime

root = tk.Tk()
root.title("MSK MRI extractor")
root.geometry("450x300") # 가로 * 세로 + x좌표 + y좌표
root.resizable(False,False) # 창 크기 변경 불가

# Title Label
label1 = Label(root, text="MSK MRI 추출기 GUI 1.1 created by H.Kim, M.D.", font = "맑은고딕 14")
label1.place(x=5,y=0)

# Label - result file name
label1 = Label(root, text="1. 결과를 저장할 파일명 입력: ", font = "맑은고딕 12")
label1.place(x=5,y=30)

# 파일 이름 받아오는 entry 생성
ety = tk.Entry(root, width=30, font = "맑은고딕 12")
ety.place(x=5,y=60)
default_file_name = (datetime.date.today() + datetime.timedelta(days=1)).strftime('%m%d')
ety.insert(END, default_file_name+".xlsx")

# Label - result file name
label1 = Label(root, text="2. 진행 상황: 아래 Extract now 버튼을 누르면 시작합니다.", font = "맑은고딕 12")
label1.place(x=5,y=90)

# progress bar
my_progress = tkinter.ttk.Progressbar(root, orient = HORIZONTAL, length=425, mode='determinate')
my_progress.place(x=5,y=120)

# print processing window
txt = Text(root, width=60, height=8, font = "맑은고딕 10")
txt.place(x=5,y=150)
txt.insert(END, "")

# Extract button 눌렀을 때 동작하는 함수
def btncmd():
    dir_path = filedialog.askopenfile(
    parent=root,initialdir=os.getcwd(),title='Please select a file',
    filetypes=(('xls files','*.xls'),('all files','*.*')))
    txt.insert(END,"다음 경로의 파일이 로드되었습니다. \n"+dir_path.name)
    df = pd.read_excel(dir_path.name, sheet_name=None)

    mri_table = pd.DataFrame()
    mri_table = pd.concat(df, ignore_index=True)

    # 필요없는 행/열들 지우기
    mri_table.drop([0, 1, 2, 3], inplace=True)
    mri_table.dropna(axis=1, inplace=True)
    mri_table.rename(columns=mri_table.iloc[0],inplace=True)
    mri_table.drop([4], inplace=True)

    # str --> int 형변환
    mri_table["NO"] = pd.to_numeric(mri_table["NO"])
    mri_table["등록번호"] = pd.to_numeric(mri_table["등록번호"])

    ## 주민 번호를 생년월일로 변환
    pre_mil = mri_table[pd.to_numeric(mri_table["주민번호"].str[7]) < 3].copy()
    pre_mil["주민번호"] = "19" + pre_mil["주민번호"]  # 주민번호 뒷자리가 1/2 인 경우 19를 붙임

    post_mil = mri_table[pd.to_numeric(mri_table["주민번호"].str[7]) > 2].copy()
    post_mil["주민번호"] = "20" + post_mil["주민번호"] # # 주민번호 뒷자리가 3/4 인 경우 20을 붙임

    mri_table = pd.merge(pre_mil,post_mil, how='outer') # 다시 합침
    mri_table = mri_table.sort_values(by='NO') # 재정렬

    mri_table["주민번호"] = mri_table["주민번호"].str[:8] # 주민번호 뒷자리 없애고 생년월일만 남김

    mri_table["주민번호"] = mri_table["주민번호"].astype('datetime64[ns]') # 변수를 날짜형으로 형변환
    mri_table = mri_table.rename(columns={'주민번호':'생년월일'}) # 열 이름을 생년월일로 변경

    # 만나이 계산
    today = datetime.date.today() # 오늘 날짜 받아오기
    age = pd.DataFrame()
    age["age"] = mri_table['생년월일'].apply(
        lambda x: today.year - x.year - ((today.month, today.day) < (x.month, x.day)))
    mri_table["생년월일"] = age["age"]
    mri_table = mri_table.rename(columns={'생년월일':'나이'})
    
    # 나이 17세 미만 환자 삭제
    mri_table = mri_table[mri_table["나이"] > 17]
    # 필요 없는 열들 삭제
    mri_table.drop(["나이","진료일자","병실"],axis=1, inplace=True)

    ## remove list 파일 읽어서 Data Frame으로 저장

    remove_list_path = "./remove_list.xlsx"
    df2 = pd.read_excel(remove_list_path, sheet_name=None, engine='openpyxl')

    remove_list_table = pd.DataFrame()
    remove_list_table = pd.concat(df2, ignore_index=True)

    ## remove list와 일치하는 처방 삭제하기
    txt.insert(END,"\n\nNon-MSK MRI 삭제 중... \n")
    table_length = len(remove_list_table)
    for i in range(table_length):
        mri_table = mri_table[~mri_table["처방명"].str.contains(remove_list_table["키워드"][i])]
        my_progress['value'] += 1/table_length*100

    # No 1부터 다시 매기기
    mri_table["NO"] = range(1, len(mri_table)+1)

    ## data frame을 openpyxl worksheet로 전환
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(mri_table, index=False, header=True):
        ws.append(r)
    ## 서식 넣기

    align_center = Alignment(horizontal='center', vertical='center', wrap_text= True)
    font_9 = Font(name='굴림', size=9, bold=False)
    border_thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    col_max = ws.max_column+2

    for x in range(1, ws.max_row + 1):
        for y in range(1, col_max):
            ws.cell(row=x, column=y).alignment = align_center
            ws.cell(row=x, column=y).font = font_9
            ws.cell(row=x, column=y).border = border_thin
    ## 셀 크기 조정
    for row in range(1,ws.max_row+1):
        ws.row_dimensions[row].height = 27.75

    for col in range(1,ws.max_column+2):
        if col == 5:
            ws.column_dimensions[get_column_letter(col)].width = 48
        elif col == 9:
            ws.column_dimensions[get_column_letter(col)].width = 48
        elif col == 10:
            ws.column_dimensions[get_column_letter(col)].width = 32
        else:
            ws.column_dimensions[get_column_letter(col)].width = 13.6

    ## 최종 결과물을 엑셀 파일로 저장
    savefilename = ety.get()
    wb.save(savefilename)
    txt.insert(END,"\n\n결과가 "+savefilename+" 으로 저장되었습니다 !!!")

# 버튼 생성
btn1 = tk.Button(root, padx=10, pady=5, text = "Extract now", command=btncmd)
btn1.place(x=177, y=260)

root.mainloop()